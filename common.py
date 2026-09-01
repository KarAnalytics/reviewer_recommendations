"""Shared helpers for the reviewer-recommendation scripts.

Handles: .env loading, talking to the local Ollama app (chat completions)
and to Ollama Cloud's web-search API, workbook auto-detection, run logging,
and parsing author names out of the Submissions sheet's "Authors" free-text
field.
"""
from __future__ import annotations

import datetime
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent


# ---------------------------------------------------------------------------
# .env loading (no external dependency -- just enough to read KEY=VALUE lines)
# ---------------------------------------------------------------------------
def load_env(env_path: Path | None = None) -> None:
    env_path = env_path or (HERE / ".env")
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_env()

OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "gemma4:31b-cloud")
OLLAMA_API_KEY = os.environ.get("OLLAMA_API_KEY", "").strip()


def _autodetect_workbook() -> Path | None:
    """Find the .xlsx to work on. Prefers WORKBOOK_PATH from .env; otherwise,
    if there's exactly one .xlsx sitting in the parent folder (next to this
    reviewer_tools/ folder), use that. Returns None if it can't tell which
    file you mean -- callers should then require --workbook explicitly.
    """
    env_val = os.environ.get("WORKBOOK_PATH")
    if env_val:
        return Path(env_val)
    candidates = [p for p in HERE.parent.glob("*.xlsx") if not p.name.startswith("~$")]
    if len(candidates) == 1:
        return candidates[0]
    return None


WORKBOOK_PATH = _autodetect_workbook()

# Set GOOGLE_SHEET_ID to switch ALL THREE scripts from the local .xlsx to a
# shared Google Sheet instead (see sheets_backend.py) -- everything else
# (WORKBOOK_PATH/--workbook) is then ignored. Also needs
# GOOGLE_SERVICE_ACCOUNT_FILE (path to your service account's JSON key
# file) or GOOGLE_SERVICE_ACCOUNT_JSON (its raw JSON content) set.
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "").strip()


def load_workbook(path: str | None):
    """Open the workbook -- a Google Sheet if GOOGLE_SHEET_ID is set
    (ignoring `path`), otherwise the local .xlsx at `path` via openpyxl.
    Use this everywhere instead of calling openpyxl.load_workbook()
    directly, so all three scripts work against either backend unchanged.
    """
    if GOOGLE_SHEET_ID:
        from sheets_backend import GoogleSheetWorkbook
        return GoogleSheetWorkbook(GOOGLE_SHEET_ID)
    import openpyxl
    return openpyxl.load_workbook(path)

# Optional context to steer the model's name-collision check and its sense
# of what "topical fit" means -- set these in .env to tailor the tool to
# your own conference. Sensible generic defaults work fine without any of
# this configured.
CONFERENCE_NAME = os.environ.get("CONFERENCE_NAME", "").strip() or "this conference"
CONFERENCE_FIELD = os.environ.get("CONFERENCE_FIELD", "").strip() or "this conference's field"


# ---------------------------------------------------------------------------
# Ollama chat (local app -- already signed in, no API key needed)
# ---------------------------------------------------------------------------
def ollama_chat(messages: list[dict], *, model: str | None = None,
                 temperature: float = 0.2, retries: int = 3) -> str:
    """Call the local Ollama app's /api/chat and return the reply text."""
    url = f"{OLLAMA_HOST}/api/chat"
    payload = {
        "model": model or OLLAMA_MODEL,
        "messages": messages,
        "stream": False,
        "options": {"temperature": temperature},
    }
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.post(url, json=payload, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            return data["message"]["content"]
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            time.sleep(min(2 ** attempt, 15))
    raise RuntimeError(
        f"Ollama chat call failed after {retries} attempts. Is the Ollama "
        f"app running? ({OLLAMA_HOST}) Last error: {last_err}"
    )


# ---------------------------------------------------------------------------
# Multi-provider chat completions -- "bring your own key".
#
# LLM_PROVIDER selects a provider (default 'ollama' -- your locally signed-
# in desktop app, unchanged from before, no key needed). LLM_API_KEY is the
# caller's OWN key for whichever provider they picked; each person running
# this (you, a co-chair, next year's co-chairs) supplies their own, with no
# shared secret to manage for this part. LLM_MODEL overrides the default
# model per provider.
#
# Providers whose wire format matches OpenAI's Chat Completions API differ
# only in base_url/default_model, so they share ONE call function
# (_openai_compatible_chat) -- add a new OpenAI-compatible provider here and
# it works immediately, no new function needed.
# ---------------------------------------------------------------------------
LLM_PROVIDER = os.environ.get("LLM_PROVIDER", "ollama").strip().lower()
LLM_API_KEY = os.environ.get("LLM_API_KEY", "").strip()
LLM_MODEL = os.environ.get("LLM_MODEL", "").strip()

_OPENAI_COMPATIBLE_PROVIDERS = {
    "openai":     {"base_url": "https://api.openai.com/v1",     "default_model": "gpt-4o-mini"},
    "groq":       {"base_url": "https://api.groq.com/openai/v1", "default_model": "llama-3.3-70b-versatile"},
    "openrouter": {"base_url": "https://openrouter.ai/api/v1",   "default_model": "openrouter/auto"},
}


def _require_llm_key(provider: str) -> str:
    if not LLM_API_KEY:
        raise RuntimeError(
            f"LLM_API_KEY is not set (required for LLM_PROVIDER={provider!r}). "
            "Set LLM_PROVIDER and LLM_API_KEY in .env."
        )
    return LLM_API_KEY


def _post_json(url: str, headers: dict, payload: dict, *, extract, provider: str,
                retries: int, timeout: int = 120) -> str:
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=timeout)
            if resp.status_code == 401:
                raise RuntimeError(f"LLM_API_KEY was rejected by {provider} (401 Unauthorized). "
                                    "Check the key.")
            resp.raise_for_status()
            return extract(resp.json())
        except RuntimeError:
            raise
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            time.sleep(min(2 ** attempt, 15))
    raise RuntimeError(f"{provider} call failed after {retries} attempts: {last_err}")


def _anthropic_chat(messages: list[dict], *, temperature: float, retries: int) -> str:
    key = _require_llm_key("anthropic")
    model = LLM_MODEL or "claude-sonnet-4-5"
    return _post_json(
        "https://api.anthropic.com/v1/messages",
        {"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        {"model": model, "max_tokens": 4096, "temperature": temperature, "messages": messages},
        extract=lambda data: "".join(
            b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"
        ),
        provider="Anthropic", retries=retries,
    )


def _openai_compatible_chat(provider: str, messages: list[dict], *, temperature: float,
                             retries: int) -> str:
    cfg = _OPENAI_COMPATIBLE_PROVIDERS[provider]
    key = _require_llm_key(provider)
    model = LLM_MODEL or cfg["default_model"]
    return _post_json(
        f"{cfg['base_url']}/chat/completions",
        {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        {"model": model, "temperature": temperature, "messages": messages},
        extract=lambda data: data["choices"][0]["message"]["content"],
        provider=provider, retries=retries,
    )


def _ollama_cloud_chat(messages: list[dict], *, temperature: float, retries: int) -> str:
    """Ollama Cloud's hosted API, authenticated with your own key -- no
    locally-signed-in desktop app needed."""
    key = _require_llm_key("ollama_cloud")
    model = LLM_MODEL or OLLAMA_MODEL
    return _post_json(
        "https://ollama.com/api/chat",
        {"Authorization": f"Bearer {key}"},
        {"model": model, "messages": messages, "stream": False,
         "options": {"temperature": temperature}},
        extract=lambda data: data["message"]["content"],
        provider="Ollama Cloud", retries=retries,
    )


def llm_chat(messages: list[dict], *, temperature: float = 0.2, retries: int = 3) -> str:
    """Provider-agnostic chat completion -- dispatches on LLM_PROVIDER.
    Defaults to 'ollama' (your locally-signed-in desktop app, no key
    needed) so existing local workflows are unaffected; set LLM_PROVIDER
    (+ LLM_API_KEY) in .env to use anthropic / openai / groq / openrouter /
    ollama_cloud instead.
    """
    if LLM_PROVIDER == "ollama":
        return ollama_chat(messages, temperature=temperature, retries=retries)
    if LLM_PROVIDER == "ollama_cloud":
        return _ollama_cloud_chat(messages, temperature=temperature, retries=retries)
    if LLM_PROVIDER == "anthropic":
        return _anthropic_chat(messages, temperature=temperature, retries=retries)
    if LLM_PROVIDER in _OPENAI_COMPATIBLE_PROVIDERS:
        return _openai_compatible_chat(LLM_PROVIDER, messages, temperature=temperature, retries=retries)
    raise RuntimeError(
        f"Unknown LLM_PROVIDER {LLM_PROVIDER!r}. Use 'ollama', 'ollama_cloud', "
        f"'anthropic', or one of {list(_OPENAI_COMPATIBLE_PROVIDERS)}."
    )


def extract_json(text: str) -> dict | list:
    """Pull the first JSON object/array out of a model reply."""
    text = text.strip()
    text = re.sub(r"^```(json)?|```$", "", text, flags=re.MULTILINE).strip()
    match = re.search(r"[\[{].*[\]}]", text, flags=re.DOTALL)
    if not match:
        raise ValueError(f"No JSON found in model reply: {text[:200]!r}")
    return json.loads(match.group(0))


# ---------------------------------------------------------------------------
# Ollama Cloud web search (needs OLLAMA_API_KEY)
# ---------------------------------------------------------------------------
MAX_RATE_LIMIT_WAIT = 30  # seconds -- never block longer than this on a single retry


def ollama_web_search(query: str, *, max_results: int = 5, retries: int = 3) -> list[dict]:
    if not OLLAMA_API_KEY:
        raise RuntimeError(
            "OLLAMA_API_KEY is not set. Copy .env.example to .env and paste "
            "a key from https://ollama.com/settings/keys into it."
        )
    url = "https://ollama.com/api/web_search"
    headers = {"Authorization": f"Bearer {OLLAMA_API_KEY}"}
    payload = {"query": query, "max_results": max_results}
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)
            if resp.status_code == 401:
                raise RuntimeError("OLLAMA_API_KEY was rejected (401 Unauthorized). "
                                    "Check the key in .env.")
            if resp.status_code == 429:
                # Rate-limited. Respect Retry-After if given, but never block
                # longer than MAX_RATE_LIMIT_WAIT on any single attempt --
                # a large/daily-quota Retry-After should surface as an error
                # (so the run can be resumed later), not hang the script.
                try:
                    retry_after = float(resp.headers.get("Retry-After", 10 * attempt))
                except ValueError:
                    retry_after = 10 * attempt
                if retry_after > MAX_RATE_LIMIT_WAIT:
                    raise RuntimeError(
                        f"Ollama Cloud web search is rate-limited and asked to wait "
                        f"{retry_after:.0f}s (likely an hourly/daily search quota) -- "
                        f"stopping here so you can resume later. Already-filled rows "
                        f"were saved; just re-run the script once the quota resets."
                    )
                last_err = "429 Too Many Requests"
                time.sleep(retry_after)
                continue
            resp.raise_for_status()
            data = resp.json()
            return data.get("results", [])
        except RuntimeError:
            raise
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            time.sleep(min(2 ** attempt, 15))
    raise RuntimeError(f"Ollama web_search failed after {retries} attempts: {last_err}")


# ---------------------------------------------------------------------------
# Person research (enrichment) -- provider-dispatched, "bring your own key".
#
# 'ollama' / 'ollama_cloud': TWO calls -- raw web_search results (Ollama
# Cloud's dedicated search endpoint; always needs OLLAMA_API_KEY for this
# part specifically, regardless of LLM_PROVIDER, since it's the only
# standalone "give me raw search results" endpoint wired up here) then a
# chat call summarizes those results into structured fields.
#
# 'anthropic' / an OpenAI-compatible provider that supports it: ONE call --
# the provider's own built-in web-search tool searches and extracts in the
# same request, authenticated with LLM_API_KEY like the other chat calls.
# This path is newer/less battle-tested than the Ollama path (which has
# been verified against 245 real reviewers) -- smoke-test with --limit 5
# before trusting it at scale.
# ---------------------------------------------------------------------------
def _research_rules() -> str:
    """Extraction rules + name-collision guardrail, shared verbatim across
    every provider's research prompt below."""
    return f"""IMPORTANT -- name-collision check: common names can match a *different*
person at the same or a different institution. Before using information,
check that it is plausibly about the SAME person: does the field of work
fit someone who co-authors or reviews for {CONFERENCE_NAME} ({CONFERENCE_FIELD})?
Information describing someone in a starkly unrelated field (e.g.
biochemistry, marine biology, high-energy physics, when this conference is
not in that space) is very likely a namesake, NOT this person, even if the
affiliation matches -- large universities have many people who share a
name. When in doubt, treat it as a different person.

From ONLY information you're confident is about this specific person,
extract:
- "position": their current job title/role (e.g. "Associate Professor of Information Systems"). If not stated (but you're still confident of the identity), use "Unknown".
- "interests": a short comma-separated list (3-8 items) of their research interests/topics. If not stated, use "".
- "website": the single best URL for this person -- prefer a personal or university faculty page, then a Google Scholar profile, then LinkedIn.

If you are not confident you found information about this specific person,
set ALL THREE fields to "Unknown"/"" -- do not mix a confident field with
information about a namesake. Do not invent information you didn't
actually find. Reply with ONLY a JSON object, no other text:
{{"position": "...", "interests": "...", "website": "..."}}"""


def _parse_research_reply(reply: str) -> dict:
    try:
        data = extract_json(reply)
    except Exception:
        return {"position": "Not found", "interests": "", "website": ""}
    return {
        "position": str(data.get("position") or "Unknown").strip()[:255],
        "interests": str(data.get("interests") or "").strip()[:500],
        "website": str(data.get("website") or "").strip()[:255],
    }


def _research_person_ollama(name: str, affiliation: str, email: str) -> dict:
    queries = [f"{name} {affiliation} research", f"{name} Google Scholar"]
    all_results = []
    for q in queries:
        try:
            all_results.extend(ollama_web_search(q, max_results=5))
        except RuntimeError:
            raise  # missing/bad API key -- not worth continuing the run
        except Exception:
            continue  # a single flaky search shouldn't kill the whole row

    if not all_results:
        return {"position": "Not found", "interests": "", "website": ""}

    snippets = "\n\n".join(
        f"URL: {r.get('url', '')}\nTitle: {r.get('title', '')}\n"
        f"Snippet: {r.get('content', r.get('snippet', ''))[:500]}"
        for r in all_results[:8]
    )

    prompt = f"""You are helping {CONFERENCE_NAME} (field: {CONFERENCE_FIELD})
identify a reviewer's profile.

Person: {name}
Stated affiliation: {affiliation or "(unknown)"}
Email: {email or "(unknown)"}

Web search results about this person:
{snippets}

{_research_rules()}"""

    return _parse_research_reply(llm_chat([{"role": "user", "content": prompt}]))


def _research_person_anthropic(name: str, affiliation: str, email: str) -> dict:
    """One call: Claude searches the web itself (server-side web_search
    tool) and extracts the answer in the same request."""
    key = _require_llm_key("anthropic")
    model = LLM_MODEL or "claude-sonnet-4-5"

    prompt = f"""You are helping {CONFERENCE_NAME} (field: {CONFERENCE_FIELD})
identify a reviewer's profile. Use web search to find current information
about this specific person before answering.

Person: {name}
Stated affiliation: {affiliation or "(unknown)"}
Email: {email or "(unknown)"}

{_research_rules()}"""

    payload = {
        "model": model,
        "max_tokens": 2048,
        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
        "messages": [{"role": "user", "content": prompt}],
    }

    def extract(data: dict) -> str:
        return "".join(
            b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"
        )

    reply = _post_json(
        "https://api.anthropic.com/v1/messages",
        {"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        payload, extract=extract, provider="Anthropic", retries=3,
    )
    return _parse_research_reply(reply)


def _research_person_openai(name: str, affiliation: str, email: str) -> dict:
    """One call via OpenAI's Responses API: the model searches the web
    itself (hosted web_search tool) and extracts the answer in the same
    request. Only plain 'openai' supports this tool today, not the other
    OpenAI-compatible providers in _OPENAI_COMPATIBLE_PROVIDERS."""
    key = _require_llm_key("openai")
    model = LLM_MODEL or "gpt-4o"

    prompt = f"""You are helping {CONFERENCE_NAME} (field: {CONFERENCE_FIELD})
identify a reviewer's profile. Use web search to find current information
about this specific person before answering.

Person: {name}
Stated affiliation: {affiliation or "(unknown)"}
Email: {email or "(unknown)"}

{_research_rules()}"""

    payload = {"model": model, "tools": [{"type": "web_search"}], "input": prompt}

    def extract(data: dict) -> str:
        chunks = []
        for item in data.get("output", []):
            if item.get("type") == "message":
                for c in item.get("content", []):
                    if c.get("type") == "output_text":
                        chunks.append(c.get("text", ""))
        return "".join(chunks) or data.get("output_text", "")

    reply = _post_json(
        "https://api.openai.com/v1/responses",
        {"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        payload, extract=extract, provider="OpenAI", retries=3,
    )
    return _parse_research_reply(reply)


def research_person(name: str, affiliation: str, email: str) -> dict:
    """Provider-dispatched: research one person, return
    {"position", "interests", "website"}. See module notes above for which
    providers do this in one call vs. two, and which need which key(s)."""
    if LLM_PROVIDER in ("ollama", "ollama_cloud"):
        return _research_person_ollama(name, affiliation, email)
    if LLM_PROVIDER == "anthropic":
        return _research_person_anthropic(name, affiliation, email)
    if LLM_PROVIDER == "openai":
        return _research_person_openai(name, affiliation, email)
    raise RuntimeError(
        f"LLM_PROVIDER={LLM_PROVIDER!r} doesn't have a research_person() path "
        "(supported: 'ollama', 'ollama_cloud', 'anthropic', 'openai'). Groq/OpenRouter "
        "work fine for suggest_reviewers.py's ranking calls (llm_chat), just not this "
        "web-search-based enrichment step."
    )


# ---------------------------------------------------------------------------
# Author-name parsing / matching
# ---------------------------------------------------------------------------
def split_authors(authors_field: str) -> list[str]:
    """Turn 'A, B, C and D' (possibly with embedded newlines) into
    ['A', 'B', 'C', 'D'].
    """
    if not authors_field:
        return []
    text = re.sub(r"\s+", " ", str(authors_field)).strip()
    text = re.sub(r"\s+and\s+", ", ", text, flags=re.IGNORECASE)
    parts = [p.strip(" .") for p in text.split(",")]
    return [p for p in parts if p]


def normalize_name(name: str) -> str:
    name = re.sub(r"\s+", " ", str(name)).strip().lower()
    name = re.sub(r"[^\w\s'-]", "", name)
    return name


def names_match(a: str, b: str) -> bool:
    na, nb = normalize_name(a), normalize_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    # last-name + first-initial fallback (handles "J. Smith" vs "John Smith")
    pa, pb = na.split(), nb.split()
    if pa and pb and pa[-1] == pb[-1] and pa[0][:1] == pb[0][:1]:
        return True
    return False


# ---------------------------------------------------------------------------
# Review-count sync -- keeps ReviewerList's No_reviews_assigned column live
# ---------------------------------------------------------------------------
REVIEW_COUNT_COLUMN = "No_reviews_assigned"
_REQUIRED_REVIEWER_SLOT_COLS = ["Reviewer 1", "Reviewer 2", "Reviewer 3"]


def sync_review_counts(wb, sub_sheet: str = "Submissions", rev_sheet: str = "ReviewerList") -> int:
    """Write/refresh a live Excel formula in ReviewerList's
    No_reviews_assigned column (creating the column if it doesn't exist
    yet) for every reviewer row. The formula COUNTIFs that reviewer's name
    across Submissions' Reviewer 1/2/3 columns, so it recalculates live in
    Excel the instant those change -- no script needs to run again unless
    new reviewer rows get added. Returns how many rows got a formula.

    Silently does nothing (returns 0) if either sheet or the Reviewer
    1/2/3 / Author columns aren't present, so callers can call this
    opportunistically without extra guarding.
    """
    if sub_sheet not in wb.sheetnames or rev_sheet not in wb.sheetnames:
        return 0
    sub_ws = wb[sub_sheet]
    rev_ws = wb[rev_sheet]

    sub_header = {str(c.value).strip(): c.column for c in sub_ws[1] if c.value}
    if any(c not in sub_header for c in _REQUIRED_REVIEWER_SLOT_COLS):
        return 0

    rev_header = {str(c.value).strip(): c.column for c in rev_ws[1] if c.value}
    if "Author" not in rev_header:
        return 0

    count_col = rev_header.get(REVIEW_COUNT_COLUMN)
    if not count_col:
        count_col = (max(rev_header.values()) if rev_header else 0) + 1
        rev_ws.cell(1, count_col).value = REVIEW_COUNT_COLUMN

    r1, r2, r3 = (get_column_letter(sub_header[c]) for c in _REQUIRED_REVIEWER_SLOT_COLS)
    author_letter = get_column_letter(rev_header["Author"])

    n = 0
    for r in range(2, rev_ws.max_row + 1):
        name = rev_ws.cell(r, rev_header["Author"]).value
        if not name or not str(name).strip():
            continue
        rev_ws.cell(r, count_col).value = (
            f"=COUNTIF({sub_sheet}!{r1}:{r1},{author_letter}{r})"
            f"+COUNTIF({sub_sheet}!{r2}:{r2},{author_letter}{r})"
            f"+COUNTIF({sub_sheet}!{r3}:{r3},{author_letter}{r})"
        )
        n += 1
    return n


# ---------------------------------------------------------------------------
# Run logging -- every run gets its own timestamped, run-numbered log file
# in logs/ (a sibling of reviewer_tools/, not inside it), while everything
# still prints to the console as normal.
# ---------------------------------------------------------------------------
LOGS_DIR = HERE.parent / "logs"


class _Tee:
    """Writes to multiple streams at once (console + log file)."""

    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for s in self.streams:
            s.write(data)

    def flush(self):
        for s in self.streams:
            s.flush()


def start_logging(script_name: str) -> Path:
    """Mirror everything this script prints (stdout + stderr) into a fresh
    timestamped, run-numbered file under logs/, in addition to the console.
    Call this once, near the top of main(). Returns the log file's path.
    """
    LOGS_DIR.mkdir(exist_ok=True)
    run_num = len(list(LOGS_DIR.glob(f"{script_name}_run*.log"))) + 1
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOGS_DIR / f"{script_name}_run{run_num:03d}_{timestamp}.log"
    log_file = open(log_path, "w", encoding="utf-8")
    sys.stdout = _Tee(sys.stdout, log_file)
    sys.stderr = _Tee(sys.stderr, log_file)
    print(f"Logging this run to {log_path}")
    return log_path
