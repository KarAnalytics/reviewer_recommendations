"""Fill in Position / Interests / Website for each person on the
ReviewerList sheet, using Ollama Cloud web search + a local Ollama model
to summarize what it finds.

Usage:
    python enrich_reviewers.py                 # fill every blank row
    python enrich_reviewers.py --limit 5        # just the first 5 (dry run)
    python enrich_reviewers.py --force          # re-look-up EVERY row, even
                                                  # ones already filled in
    python enrich_reviewers.py --only "Pei Siang Goh,Xiaobai Li"  # redo just these people
    python enrich_reviewers.py --workbook path\to\file.xlsx

Only rows where Position, Interests, and Website are ALL blank get looked
up (unless --force is given), so re-running after adding a few new names
to ReviewerList only does work for the new people.

Requires OLLAMA_API_KEY in .env (see .env.example) and the Ollama app
running locally (for the summarizing chat calls).
"""
from __future__ import annotations

import argparse
import sys
import time

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

from common import WORKBOOK_PATH, CONFERENCE_NAME, CONFERENCE_FIELD, ollama_chat, ollama_web_search, extract_json

SHEET = "ReviewerList"
SAVE_EVERY = 1  # autosave cadence, so a crash mid-run doesn't lose progress


def find_columns(ws) -> dict[str, int]:
    header = {}
    for cell in ws[1]:
        if cell.value:
            header[str(cell.value).strip()] = cell.column
    required = ["Author", "Affiliation", "Email", "Position", "Interests", "Website"]
    missing = [c for c in required if c not in header]
    if missing:
        raise SystemExit(f"ReviewerList is missing expected column(s): {missing}")
    return header


def lookup_reviewer(name: str, affiliation: str, email: str) -> dict:
    """Web-search for a person, then have the model extract structured info."""
    queries = [f"{name} {affiliation} research", f"{name} Google Scholar"]
    all_results = []
    for q in queries:
        try:
            all_results.extend(ollama_web_search(q, max_results=5))
        except RuntimeError:
            raise  # missing/bad API key -- not worth continuing the run
        except Exception:
            continue  # a single flaky search shouldn't kill the whole row

    if not all_results:
        return {"position": "Not found", "interests": "", "website": ""}

    snippets = "\n\n".join(
        f"URL: {r.get('url','')}\nTitle: {r.get('title','')}\n"
        f"Snippet: {r.get('content', r.get('snippet',''))[:500]}"
        for r in all_results[:8]
    )

    prompt = f"""You are helping {CONFERENCE_NAME} (field: {CONFERENCE_FIELD})
identify a reviewer's profile.

Person: {name}
Stated affiliation: {affiliation or "(unknown)"}
Email: {email or "(unknown)"}

Web search results about this person:
{snippets}

IMPORTANT -- name-collision check: common names can match a *different*
person at the same or a different institution. Before using a result,
check that it is plausibly the SAME person: does the field of work fit
someone who co-authors or reviews for {CONFERENCE_NAME} ({CONFERENCE_FIELD})?
A result describing someone in a starkly unrelated field (e.g. biochemistry,
marine biology, high-energy physics, when this conference is not in that
space) is very likely a namesake, NOT this person, even if the affiliation
string matches -- large universities have many people who share a name.
When in doubt, treat it as a different person.

From ONLY search results you're confident are about this specific person,
extract:
- "position": their current job title/role (e.g. "Associate Professor of Information Systems"). If not stated (but you're still confident of the identity), use "Unknown".
- "interests": a short comma-separated list (3-8 items) of their research interests/topics. If not stated, use "".
- "website": the single best URL for this person -- prefer a personal or university faculty page, then a Google Scholar profile, then LinkedIn.

If you are not confident ANY result is about this specific person, set ALL
THREE fields to "Unknown"/"" -- do not mix a confident field with fields
drawn from a namesake's results. Do not invent information not supported
by the search results. Reply with ONLY a JSON object, no other text:
{{"position": "...", "interests": "...", "website": "..."}}"""

    reply = ollama_chat([{"role": "user", "content": prompt}])
    try:
        data = extract_json(reply)
    except Exception:
        return {"position": "Not found", "interests": "", "website": ""}

    return {
        "position": str(data.get("position") or "Unknown").strip()[:255],
        "interests": str(data.get("interests") or "").strip()[:500],
        "website": str(data.get("website") or "").strip()[:255],
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", default=str(WORKBOOK_PATH) if WORKBOOK_PATH else None,
                     required=WORKBOOK_PATH is None,
                     help="path to the .xlsx (auto-detected if there's exactly one "
                          "next to reviewer_tools/; otherwise required, or set "
                          "WORKBOOK_PATH in .env)")
    ap.add_argument("--limit", type=int, default=None, help="only process the first N eligible rows")
    ap.add_argument("--force", action="store_true", help="re-look-up rows even if already filled in")
    ap.add_argument("--only", default=None,
                     help="comma-separated exact reviewer name(s) to (re-)look-up, "
                          "ignoring the already-filled-in check")
    args = ap.parse_args()
    only_names = None
    if args.only:
        only_names = {n.strip().lower() for n in args.only.split(",") if n.strip()}

    wb = openpyxl.load_workbook(args.workbook)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"No '{SHEET}' sheet in {args.workbook}")
    ws = wb[SHEET]
    col = find_columns(ws)

    todo = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col["Author"]).value
        if not name or not str(name).strip():
            continue
        if only_names is not None:
            if str(name).strip().lower() in only_names:
                todo.append(r)
            continue
        pos = ws.cell(r, col["Position"]).value
        interests = ws.cell(r, col["Interests"]).value
        website = ws.cell(r, col["Website"]).value
        already_done = any(v not in (None, "") for v in (pos, interests, website))
        if already_done and not args.force:
            continue
        todo.append(r)

    if args.limit:
        todo = todo[: args.limit]

    if not todo:
        print("Nothing to do -- every reviewer already has Position/Interests/Website filled in.")
        print("(Pass --force to re-look-up everyone anyway.)")
        return

    print(f"Looking up {len(todo)} reviewer(s)...")
    t0 = time.time()
    for i, r in enumerate(todo, 1):
        name = str(ws.cell(r, col["Author"]).value).strip()
        affiliation = ws.cell(r, col["Affiliation"]).value or ""
        email = ws.cell(r, col["Email"]).value or ""

        print(f"[{i}/{len(todo)}] {name} ...", end=" ", flush=True)
        try:
            info = lookup_reviewer(name, str(affiliation), str(email))
        except RuntimeError as exc:
            print(f"\nFATAL: {exc}")
            sys.exit(1)
        except Exception as exc:  # noqa: BLE001
            print(f"error ({exc}), skipping")
            continue

        ws.cell(r, col["Position"]).value = info["position"]
        ws.cell(r, col["Interests"]).value = info["interests"]
        ws.cell(r, col["Website"]).value = info["website"]
        print(info["position"] or "-")

        if i % SAVE_EVERY == 0:
            wb.save(args.workbook)

    wb.save(args.workbook)
    elapsed = time.time() - t0
    print(f"Done. Saved {args.workbook} ({elapsed:.0f}s, {elapsed/max(len(todo),1):.1f}s/reviewer avg).")


if __name__ == "__main__":
    main()
