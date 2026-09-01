"""Populate the AISuggestedReviewers column on the Submissions sheet: for
each paper, ask the local Ollama model to pick the best-matching reviewers
from ReviewerList, automatically excluding the paper's own authors (and
anyone already entered as Reviewer 1/2/3 for that paper).

A reviewer is also dropped from consideration once they've been suggested
--max-per-reviewer times (default 5) across the whole run, so suggestions
spread across the pool instead of a few well-matched people getting
suggested for nearly every paper. This counts suggestions already sitting
in AISuggestedReviewers from earlier runs too, so the cap holds even
across repeated runs.

Among reasonably-fitting candidates, ranking primarily favors early-career
reviewers (PhD students/postdocs/assistant-associate professors) over
senior leadership/administrative roles (dean/chair/director/etc.), who
tend to respond more slowly -- seniors are used as a fallback, not a
coequal option, unless clearly the best topical fit. A same-seniority
secondary tiebreak then prefers whoever currently has fewer actual
Reviewer 1/2/3 assignments, to spread workload.

Usage:
    python suggest_reviewers.py                  # fill blank AISuggestedReviewers cells only
    python suggest_reviewers.py --force           # recompute every paper's suggestions
    python suggest_reviewers.py --top-n 5         # how many reviewers to suggest (default 5)
    python suggest_reviewers.py --max-per-reviewer 5  # suggestion cap per reviewer (default 5)
    python suggest_reviewers.py --workbook path\to\file.xlsx

Run this any time you add new papers (or re-run enrich_reviewers.py) --
it only (re)computes rows that need it, so it's safe/cheap to re-run.

Requires the Ollama app running locally with the model set in .env
(OLLAMA_MODEL, default gemma4:31b-cloud). No OLLAMA_API_KEY needed for
this script -- that's only for enrich_reviewers.py's web search.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

from common import WORKBOOK_PATH, CONFERENCE_NAME, ollama_chat, extract_json, split_authors, names_match, start_logging, sync_review_counts

SUB_SHEET = "Submissions"
REV_SHEET = "ReviewerList"
SAVE_EVERY = 5


def find_columns(ws, required: list[str]) -> dict[str, int]:
    header = {}
    for cell in ws[1]:
        if cell.value:
            header[str(cell.value).strip()] = cell.column
    missing = [c for c in required if c not in header]
    if missing:
        raise SystemExit(f"'{ws.title}' is missing expected column(s): {missing}")
    return header


def load_reviewer_pool(ws, col: dict[str, int]) -> list[dict]:
    pool = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col["Author"]).value
        if not name or not str(name).strip():
            continue
        pool.append({
            "name": str(name).strip(),
            "position": str(ws.cell(r, col["Position"]).value or "").strip(),
            "interests": str(ws.cell(r, col["Interests"]).value or "").strip(),
            "affiliation": str(ws.cell(r, col["Affiliation"]).value or "").strip(),
        })
    return pool


def build_own_submission_index(sub_ws, sub_col: dict[str, int]) -> list[dict]:
    """Every submitted paper with its row + authors, so we can tell a
    candidate reviewer "you yourself submitted a paper on X" -- a much
    stronger interest signal than a bio blurb, and it's pulled fresh from
    Submissions each run rather than baked into ReviewerList.
    """
    papers = []
    for r in range(2, sub_ws.max_row + 1):
        title = sub_ws.cell(r, sub_col["Title"]).value
        if not title or not str(title).strip():
            continue
        if not sub_ws.cell(r, sub_col["paper"]).value:
            continue
        papers.append({
            "row": r,
            "title": str(title).strip(),
            "keywords": str(sub_ws.cell(r, sub_col["Keywords"]).value or "").strip(),
            "authors": split_authors(str(sub_ws.cell(r, sub_col["Authors"]).value or "")),
        })
    return papers


def build_assignment_counts(sub_ws, sub_col: dict[str, int], pool: list[dict]) -> dict[str, int]:
    """How many papers each reviewer is *actually* assigned to (Reviewer
    1/2/3 filled in) right now -- their real current workload, as opposed
    to how many times they've been *suggested*. Used as a ranking
    tiebreak: prefer whoever has fewer assignments so far.
    """
    assigned_names = []
    for r in range(2, sub_ws.max_row + 1):
        if not sub_ws.cell(r, sub_col["Title"]).value:
            continue
        if not sub_ws.cell(r, sub_col["paper"]).value:
            continue
        for c in ("Reviewer 1", "Reviewer 2", "Reviewer 3"):
            val = sub_ws.cell(r, sub_col[c]).value
            if val and str(val).strip():
                assigned_names.append(str(val).strip())

    counts = {c["name"]: 0 for c in pool}
    for assigned in assigned_names:
        for name in counts:
            if names_match(name, assigned):
                counts[name] += 1
                break
    return counts


_SUGGESTION_LINE_RE = re.compile(r"^\s*\d+\.\s*(.+?)(?:\s+--\s+.*)?\s*$")


def parse_suggested_names(cell_text) -> list[str]:
    """Pull reviewer names back out of an AISuggestedReviewers cell written
    by this script (numbered list, optionally "-- reason" per line).
    """
    if not cell_text:
        return []
    names = []
    for line in str(cell_text).splitlines():
        m = _SUGGESTION_LINE_RE.match(line)
        if m:
            names.append(m.group(1).strip())
    return names


def own_papers_for(name: str, all_papers: list[dict], exclude_row: int) -> list[dict]:
    return [
        p for p in all_papers
        if p["row"] != exclude_row and any(names_match(name, a) for a in p["authors"])
    ]


def pick_reviewers(title: str, keywords: str, abstract: str,
                    candidates: list[dict], top_n: int) -> list[str]:
    if not candidates:
        return []

    lines = []
    for i, c in enumerate(candidates):
        line = (f"{i+1}. {c['name']} -- {c['position'] or 'position unknown'} "
                f"({c['affiliation'] or 'affiliation unknown'}). "
                f"Interests: {c['interests'] or 'unknown'}. "
                f"Currently assigned to review {c.get('assigned_count', 0)} paper(s).")
        if c.get("own_papers"):
            own = "; ".join(
                f"\"{p['title']}\" (keywords: {p['keywords'] or 'n/a'})"
                for p in c["own_papers"]
            )
            line += f" | Also a {CONFERENCE_NAME} author, on: {own}"
        lines.append(line)
    listing = "\n".join(lines)

    prompt = f"""You are helping a conference track co-chair assign peer reviewers.

PAPER
Title: {title}
Keywords: {keywords}
Abstract: {abstract}

CANDIDATE REVIEWERS (numbered list; you may ONLY choose from this list):
{listing}

First, narrow to candidates who are a *reasonable* topical fit for this
paper -- their stated Interests (or, if those are unknown/thin, a related
paper they themselves authored for {CONFERENCE_NAME}, shown below) should
plausibly connect to this paper's topic. Don't include someone with no
plausible connection just to satisfy a preference below -- fit is a
requirement, not just a preference.

Among that reasonably-fitting group, rank using these priorities, in order:
1. PRIMARY: prefer early-career / junior reviewers -- PhD students,
   postdocs, and assistant/associate professors -- over senior
   leadership/administrative roles (dean, vice dean, department
   head/chair, director, provost, president, CEO, or a
   named/distinguished/endowed chair professorship). Junior researchers
   typically respond to and complete review requests fast; senior people
   in those roles are usually stretched thin with editorial boards and
   other service commitments. Only include a senior/leadership candidate
   when they are clearly and substantially the best topical match and no
   reasonably-fitting junior/early-career candidate covers this topic --
   seniors are a fallback, not a coequal option.
2. SECONDARY, among candidates of similar seniority tier and topical fit:
   prefer whoever is currently assigned to review fewer papers (see
   "Currently assigned to N paper(s)" above), to spread the actual review
   workload instead of piling more onto people already reviewing several.
3. A clearly superior topical match still wins over a weaker one even if
   it means picking someone more senior or busier -- these priorities
   break ties and near-ties, they don't override an obviously better fit.

Reply with ONLY a JSON array of exactly {top_n} objects (fewer only if the
list has fewer reasonably-fitting candidates), ranked best-fit first:
[{{"name": "<exact name as listed above>", "reason": "<one short clause, <12 words>"}}]"""

    reply = ollama_chat([{"role": "user", "content": prompt}])
    try:
        data = extract_json(reply)
    except Exception:
        return []

    by_name = {c["name"].lower(): c["name"] for c in candidates}
    picks = []
    for item in data:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        canonical = by_name.get(name.lower())
        if not canonical:
            # loose fallback match in case the model paraphrased slightly
            for cand in candidates:
                if names_match(name, cand["name"]):
                    canonical = cand["name"]
                    break
        if canonical and canonical not in [p[0] for p in picks]:
            picks.append((canonical, str(item.get("reason", "")).strip()))

    return picks[:top_n]


def main() -> None:
    start_logging("suggest_reviewers")
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", default=str(WORKBOOK_PATH) if WORKBOOK_PATH else None,
                     required=WORKBOOK_PATH is None,
                     help="path to the .xlsx (auto-detected if there's exactly one "
                          "next to reviewer_tools/; otherwise required, or set "
                          "WORKBOOK_PATH in .env)")
    ap.add_argument("--top-n", type=int, default=5)
    ap.add_argument("--force", action="store_true", help="recompute every paper, not just blank ones")
    ap.add_argument("--limit", type=int, default=None, help="only process the first N eligible papers")
    ap.add_argument("--max-per-reviewer", type=int, default=5,
                     help="drop a reviewer from consideration once they've been suggested "
                          "this many times across the run (default 5); 0 disables the cap")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    for sheet in (SUB_SHEET, REV_SHEET):
        if sheet not in wb.sheetnames:
            raise SystemExit(f"No '{sheet}' sheet in {args.workbook}")

    sub_ws = wb[SUB_SHEET]
    rev_ws = wb[REV_SHEET]

    sub_col = find_columns(sub_ws, [
        "Authors", "Title", "paper", "Keywords", "Abstract",
        "Reviewer 1", "Reviewer 2", "Reviewer 3", "AISuggestedReviewers",
    ])
    rev_col = find_columns(rev_ws, ["Author", "Affiliation", "Position", "Interests"])

    # Keep ReviewerList's No_reviews_assigned formulas in sync (adds the
    # column if missing, fills it down into any newly-added reviewer rows).
    sync_review_counts(wb)

    pool = load_reviewer_pool(rev_ws, rev_col)
    if not pool:
        raise SystemExit("ReviewerList has no reviewers to suggest from.")

    all_papers = build_own_submission_index(sub_ws, sub_col)
    assignment_counts = build_assignment_counts(sub_ws, sub_col, pool)

    todo = []
    for r in range(2, sub_ws.max_row + 1):
        title = sub_ws.cell(r, sub_col["Title"]).value
        if not title or not str(title).strip():
            continue
        paper_flag = sub_ws.cell(r, sub_col["paper"]).value
        if not paper_flag:
            continue  # withdrawn / not a real submission
        existing = sub_ws.cell(r, sub_col["AISuggestedReviewers"]).value
        if existing and not args.force:
            continue
        todo.append(r)

    if args.limit:
        todo = todo[: args.limit]

    if not todo:
        print("Nothing to do -- every submitted paper already has AISuggestedReviewers filled in.")
        print("(Pass --force to recompute everyone anyway.)")
        return

    # Seed the per-reviewer suggestion cap from rows that WON'T be
    # recomputed this run, so the cap holds correctly across repeated runs
    # (and across --force, which regenerates the rows in `todo` from
    # scratch rather than double-counting their old suggestions).
    todo_set = set(todo)
    suggestion_counts: Counter[str] = Counter()
    if args.max_per_reviewer > 0:
        pool_names = {c["name"].lower(): c["name"] for c in pool}
        for r in range(2, sub_ws.max_row + 1):
            if r in todo_set:
                continue
            for name in parse_suggested_names(sub_ws.cell(r, sub_col["AISuggestedReviewers"]).value):
                canonical = pool_names.get(name.strip().lower(), name.strip())
                suggestion_counts[canonical] += 1

    print(f"Suggesting reviewers for {len(todo)} paper(s)...")
    t0 = time.time()
    for i, r in enumerate(todo, 1):
        title = str(sub_ws.cell(r, sub_col["Title"]).value or "").strip()
        keywords = str(sub_ws.cell(r, sub_col["Keywords"]).value or "").strip()
        abstract = str(sub_ws.cell(r, sub_col["Abstract"]).value or "").strip()
        authors_field = sub_ws.cell(r, sub_col["Authors"]).value or ""
        paper_authors = split_authors(str(authors_field))

        already_assigned = [
            str(sub_ws.cell(r, sub_col[c]).value).strip()
            for c in ("Reviewer 1", "Reviewer 2", "Reviewer 3")
            if sub_ws.cell(r, sub_col[c]).value
        ]

        excluded_names = paper_authors + already_assigned
        candidates = []
        for c in pool:
            if any(names_match(c["name"], ex) for ex in excluded_names):
                continue
            if args.max_per_reviewer > 0 and suggestion_counts[c["name"]] >= args.max_per_reviewer:
                continue
            c = dict(c, own_papers=own_papers_for(c["name"], all_papers, exclude_row=r),
                     assigned_count=assignment_counts.get(c["name"], 0))
            candidates.append(c)

        print(f"[{i}/{len(todo)}] #{sub_ws.cell(r, sub_col['Title']).value and r} {title[:60]!r} "
              f"({len(candidates)} eligible reviewers) ...", end=" ", flush=True)

        try:
            picks = pick_reviewers(title, keywords, abstract, candidates, args.top_n)
        except Exception as exc:  # noqa: BLE001
            print(f"error ({exc}), skipping")
            continue

        if not picks:
            print("no suggestions returned")
            continue

        cell_text = "\n".join(
            f"{j}. {name}" + (f" -- {reason}" if reason else "")
            for j, (name, reason) in enumerate(picks, 1)
        )
        sub_ws.cell(r, sub_col["AISuggestedReviewers"]).value = cell_text
        for name, _reason in picks:
            suggestion_counts[name] += 1
        print(f"{len(picks)} suggested")

        if i % SAVE_EVERY == 0:
            wb.save(args.workbook)

    wb.save(args.workbook)
    elapsed = time.time() - t0
    print(f"Done. Saved {args.workbook} ({elapsed:.0f}s, {elapsed/max(len(todo),1):.1f}s/paper avg).")


if __name__ == "__main__":
    main()
